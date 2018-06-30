from openpyxl import load_workbook
import pickle
import datetime
import re

# Static data paths
JOB_DATA = './data/professions.xlsx'
COMPETENCE_RELEVANCE = './data/dspp_to_dscf.xlsx'

# static position level mapping
position_level = {
    # 0 - 2 years experience
    'entry': ['intern', 'junior', 'entry', 'staff role'],
    # 2 - 5 years experience
    'intermediate': ['intermediate'],
    # 5 - 8 years experience
    'senior': ['senior', 'sr'],
    # 5 - 8 years experience
    'principal': ['principal'],
    # > 8 years experience
    'lead': ['lead', 'chief']
}

competence_list = ['DSDA01', 'DSDA02','DSDA03', 'DSDA04', 'DSDA05', 'DSDA06', 'DSENG01', 'DSENG02',
    'DSENG03', 'DSENG04', 'DSENG05', 'DSENG06', 'DSDM01', 'DSDM02', 'DSDM03', 'DSDM04', 'DSDM05',
    'DSDM06', 'DSRM01', 'DSRM02', 'DSRM03', 'DSRM04', 'DSRM05', 'DSRM06', 'DSDK01', 'DSDK02', 'DSDK03',
    'DSDK04', 'DSDK05', 'DSDK06']

# Regex for finding amount of experience
# Find a year duration: e.g: 2005 to 2012, 2012 - 2015, 2015-2017, 2017 - Current
year_duration = r'(?i)((20|19)(\d{2}))(( - )|(-)|(to)|( to ))(((20|19)(\d{2}))|now|current|present)'

# Find a month year duration
months_short = r'(jan)|(feb)|(mar)|(apr)|(may)|(jun)|(jul)|(aug)|(sep)|(oct)|(nov)|(dec)'
months_long = r'(january)|(february)|(march)|(april)|(may)|(june)|(july)|(august)|(september)|(october)|(november)|(december)'
month = r'('+months_short+r'|'+months_long+r')'
year = r'((20|19)(\d{2})|(\d{2}))'
month_year_duration = r'(?i)' + month + r' ' + year + r'(-| - | to |to)' + month + r' ' + year

# Find a full date duration: e.g: 05/09/1996 - 3/9/2005 (dd/mm/yyyy)
full_date_duration = r'(?i)(0?[1-9]|[12][0-9]|3[01])(/|-)(0?[1-9]|1[0-2])(/|-)((20|19)(\d{2})|(\d{2}))(( - )|(-)|(to)|( to ))(((0?[1-9]|[12][0-9]|3[01])(/|-)(0?[1-9]|1[0-2])(/|-)((20|19)(\d{2})|(\d{2})))|now|current|present)'

# Example CV's to run program without API usage
EXAMPLE_CV_1 = 'Hadoop Developer - 3 Year Resume Sample Ivy Haddington • Alpharetta, GA • (123) 456-7891 • ihaddington@email.com SUMMARY Hortonworks-certified Hadoop Developer with 3+ years of experience installing, configuring, and leveraging the Hadoop ecosystem to glean meaningful insights from semi-structured and unstructured data. EDUCATION LONGFORD TECH Aug \'10 - May \'14 • Bachelor of Science in Computer Engineering EXPERIENCE RETAIL OCEAN Hadoop Developer Sep \'14 - Current • Implemented Hadoop data pipeline to identify customer behavioral patterns, improving UX on e- commerce website • Develop MapReduce jobs in Java for log analysis, analytics, and data cleaning • Perform big data processing using Hadoop, MapReduce, Sqoop, Oozie, and Impala • Import data from MySQL to HDFS, using Sqoop to load data • Developed and designed a 10-node Hadoop cluster for sample data analysis • Regularly tune performance of Hive and Pig queries to improve data processing and retrieving • Run Hadoop streaming jobs to process terabytes of XML data • Create visualizations and reports for the business intelligence team, using Tableau CRANE & JENKINS Hadoop Developer Intern Mar \'14 - May \'14 • Analyzed datasets using Pig, Hive, MapReduce, and Sqoop to recommend business improvements • Setup, installed, and monitored 3-node enterprise Hadoop cluster on Ubuntu Linux • Analyzed and interpreted transaction behaviors and clickstream data with Hadoop and HDP to predict what customers might buy in the future SKILLS • Hadoop big data ecosystems (MapReduce, HDFS, HBase, Zookeeper, Hive, Pig, Sqoop, Cassandra, Oozie, Talend) mailto:ihaddington@email.com • Java, C/C++, Python, Bash • Data modeling, analysis, and mining • Machine learning'
EXAMPLE_CV_2 = 'Hadoop Developer - 15 Year Resume Sample Cody Fredrickson • Boston, MA • (123) 456-7891 • cfredrickson@email.com SUMMARY Principal Hadoop Developer with 15+ years of experience building scalable, distributed data solutions with 80TB+ of data and driving business improvements with innovative Hadoop and BI tools. EDUCATION GREEN VALLEY STATE Aug \'98 May \'02 Master of Science in Computer Science EXPERIENCE RIVER TECH Principal Hadoop Developer Jul \'15 - Current • Leverage Hadoop and HDP to analyze massive amounts of clickstream data and identify the most efficient path for customers making an online purchase • Analyze Hadoop clusters using big data analytic tools including Pig, Hive, and MapReduce • Conduct in-depth research on Hive to analyze partitioned and bucketed data CRANE & JENKINS Senior Hadoop Developer Jan \'13 - Jun \'15 • Developed Oozie workflow to automate the loading of data into HDFS and Pig for data pre- processing • Architected 60-node Hadoop clusters with CDH4.4 on CentOS • Successfully implemented Cloudera on a 30-node cluster TRADELOT Hadoop Developer Mar \'02 - Dec \'12 • Leveraged Sqoop to import data from RDBMS into HDFS • Developed ETL framework using Python and Hive (including daily runs, error handling, and logging) to glean useful data and improve vendor negotiations • Performed cleaning and filtering on imported data using Hive and MapReduce SKILLS • Hadoop ecosystem (HDFS, Spark, Sqoop, Flume, Hive, Impala, MapReduce, Sentry, Navigator) • Hadoop data ingestion using ETL tools (MapReduce, Spark, Blaze) • Java, J2EE, C/C++, .NET'
EXAMPLE_CV_3 = 'CV_DataScientist_HeZhang CV for Data Scientist HE ZHANG Data Scientist, PhD in Machine Learning Address: Innopoli 2, FI-02150, Espoo, Finland Tel: +358-505188888 Email: klarke4001@gmail.com Born: 19.08.1981, Changchun, P. R. China KEY COMPETENCIES AND STRENGTHS • Over 7 years research and working experience in Machine Learning and Data Mining field. • Strong data analytical and programming skills especially with Matlab and Python. • Excellent English writing and oral presentation skills. • Strong team-work spirit with experience of working in highly international environments for years. • Native Mandarin speaker with Permanent Finnish Resident and Working Permit. WORKING & RESEARCH EXPERIENCE 2014 - 2015 Data Analyst at Verto Analytics Inc. (Area: data analytics and image recognition) I am working on versatile projects at Verto Analytics Inc. - a Finnish Pioneer in Digital Media Research and Measurement Industry. My responsibilities include: 1) developing and implementing machine learning algorithms for mobile-end App image recognition; 2) collaborating with marketing professionals for writing market insights reports; 3) data quality assurance, data cleaning and curation, data visualisation, and data production. 2011 - 2014 Nonnegative Learning for Data Clustering (Area: algorithms and optimisations) I designed several Machine learning algorithms using matrix factorisation models to better detect groups or clusters in various data sets. The algorithms can be directly applied for, e.g., Recommendation Systems and Market Segmentation. I published the results in 6 scientific journals and papers. 2011 - 2014 Understanding the Emotional Impacts of Images (Area: image processing) I developed several image processing methods to predict emotional impacts of artistic images. The methods can improve the performance of Affective Image Classification and Retrieval systems. I published the results in 4 scientific journals and papers. 2008 - 2010 PinView - A Proactive Personal Information Navigator (Area: multimedia retrieval) I developed a Gaze-and-Speech-enhanced Content-Based Image Retrieval system that can infer the user\'s search interests based on his or her feedbacks such as eye tracking data. I also implemented a client-side browser extension using JavaScript and managed to publish the results in 2 scientific conferences. 2005 - 2007 Research Assistant in the Multimedia Laboratory, Jilin University, China I developed matrix transformation techniques for colour image and video compression. CV: He Zhang, +358-50-5188888, klarke4001@gmail.com Page 1/3 mailto:klarke4001@gmail.com Now mailto:klarke4001@gmail.com LANGUAGE & IT SKILLS English: Excellent in Writing and Speaking. IELTS Score (2006): 7.5/9. I also have English-Chinese Translator & Interpreter experience with certificate issued by China HR Ministry. Finnish: Basic. I received full scores in 4 consecutive Aalto University Finnish Exams 2006-2007. Chinese: Native. Programming: Matlab, Python, SQL, Perl, JavaScript, C / C++, LaTeX (for document writing) POSITIONS OF TRUST 2014 Programme Committee Member in 2014 International Conference on Artificial Neural Networks (ICANN), Hamburg, Germany. 2014 Reviewer for Scientific Journals, e.g., IEEE Transactions on Neural Networks and Learning Systems, Information Sciences, Neurocomputing, Journal of Optical Engineering. 2013 Membership in European Neural Networks Society (ENNS) REFEREES Professor Erkki Oja, PhD Supervisor Email: erkki.oja@aalto.fi Department of Computer Science, Aalto University School of Science, Espoo, Finland Professor Timo Honkela, Research Collaborator Email: timo.honkela@helsinki.fi Department of Language, University of Helsinki, Helsinki, Finland Senior Scientist Jorma Laaksonen, (former) PhD Instructor Email: jorma.laaksonen@aalto.fi Department of Computer Science, Aalto University School of Science, Espoo, Finland EDUCATION 2008 - 2014 Doctor of Science, Aalto University School of Science, Finland Research areas: Machine Learning, Data Mining, and Image Processing Minor: Signal Processing for Tele-communications. PhD Advisor: Prof. Erkki Oja 2004 - 2007 Master of Science, Jilin University, China Major: Information and Communication Systems Master Thesis: Matrix Transformation Techniques for Color Image Compression 2000 - 2004 Bachelor of Engineering, Jilin University, China Major: Communication Engineering CV: He Zhang, +358-50-5188888, klarke4001@gmail.com Page 2/3 mailto:klarke4001@gmail.com mailto:klarke4001@gmail.com PUBLICATION LIST Journal Articles 1. He Zhang, Zhirong Yang, and Erkki Oja. Improving Cluster Analysis By Co-initialisations. Pattern Recognition Letters, 45: 71-77, 2014. 2. He Zhang, Zhirong Yang, and Erkki Oja. Adaptive Multiplicative Updates for Quadratic Nonnegative Matrix Factorisation. Neurocomputing, 134: 206-213, 2014. 3. He Zhang, Mehmet Gönen, Zhirong Yang, and Erkki Oja. Understanding Emotional Impact of Images Using Bayesian Multiple Kernel Learning. Neurocomputing, 165: 3-13, 2015. Conference Papers 4. He Zhang, Mehmet Gönen, Zhirong Yang, and Erkki Oja. Predicting Emotional States of Images Using Bayesian Multiple Kernel Learning. In Proceedings of the 20th International Conference on Neural Information Processing (ICONIP), Daegu, South Korea, 2013. Oral presentation. 5. He Zhang, Zhirong Yang, Mehmet Gönen, Markus Koskela, Jorma Laaksonen, Timo Honkela, and Erkki Oja. Affective Abstract Image Classification and Retrieval Using Multiple Kernel Learning. ICONIP 2013, Daegu, South Korea, 2013. Oral presentation. 6. He Zhang, Zhirong Yang, and Erkki Oja. Adaptive Multiplicative Updates for Projective Nonnegative Matrix Factorisation. ICONIP 2012, Doha, Qatar, 2012. Oral presentation. 7. Zhirong Yang, He Zhang, and Erkki Oja. Online Projective Nonnegative Matrix Factorisation for Large Datasets. ICONIP 2012, Doha, Qatar, 2012. Oral presentation. 8. He Zhang, Tele Hao, Zhirong Yang, and Erkki Oja. Pairwise Clustering with t-PLSI. In Proceedings of the 22nd International Conference on Artificial Neural Networks (ICANN), Lausanne, Switzerland, 2012. Travel Grant Award. 9. He Zhang, Mats Sjöberg, Jorma Laaksonen, and Erkki Oja. A Multimodal Information Collector for Content-Based Image Retrieval System. ICONIP 2011, Shanghai, China, 2011. Oral presentation. 10. He Zhang, Eimontas Augilius, Timo Honkela, Jorma Laaksonen et al. Analysing Emotional Semantics of Abstract Art Using Low-Level Image Features. In Proceedings of the 10th International Conference on Advances in Intelligent Data Analysis (IDA), Porto, Portugal, 2011. Oral presentation. 11. He Zhang, Teemu Ruokolainen, Jorma Laaksonen, Christina Hochleitner, and Rudolf Traunmüller. Gaze and Speech-Enhanced Content-Based Image Retrieval in Image Tagging. ICANN 2011, Espoo, Finland, 2011. Poster presentation. 12. Zhirong Yang, He Zhang, Zhijian Yuan, and Erkki Oja. Kullback-Leibler Divergence for Nonnegative Matrix Factorisation. ICANN 2011, Espoo, Finland, 2011. Oral presentation. Technical Reports 13. He Zhang, Markus Koskela, and Jorma Laaksonen. Report on Forms of Enriched Relevance Feedback. Technical Report TKK-ICS-R10, Helsinki University of Technology, Department of Information and Computer Science. Presented at PinView meeting, University College London, 2008. 14. He Zhang, Mats Sjöberg, and Jorma Laaksonen. Browser Extension for Pointer Track Feedback. PinView Deliverables. Presented at PinView meeting, Leoben University, Austria, 2008. 15. Christina Hochleitner, Rudolf Traunmuller, Teemu Ruokolainen, and He Zhang. Archiving with Supported Tagging. PinView Deliverables. Presented in Celum Company, Austria, 2011. CV: He Zhang, +358-50-5188888, klarke4001@gmail.com Page 3/3 mailto:klarke4001@gmail.com mailto:klarke4001@gmail.com'
EXAMPLE_CV_4 = 'https://www.livecareer.com/build-resume/select-resume Data Scientist Jason Caldwell 1141 Briarcliffe Road, Portland, OR 11111 T: 555-485-5897 E: jasoncaldwell@anymail Professional Summary Experienced and driven data scientist with eight years of experience in the industry. Strong background in computer programming language, and knowledge of various types of databases. Solid skills in mathematics, statistics and algorithms. Commitment to providing support and essential information about trends to companies in a variety of industries. Work Experience Data Scientist July 2014 - present • Research and transform information from raw data into an easily understood analysis that identifies trends and insights for the organization. • Pinpoint a set of variables to evaluate and work with when deciding on the range of analysis and scope of information sought. • Use a variety of sources inside and outside of the company to collect, aggregate and analyze data, using corresponding Big Data platforms and NoSQL databases Hbase, MongoDB, Azure DocumentDB • Work with business oriented Big Data Analytics platforms, Hadoop tools, Tableau visualisation tools Data Scientist August 2011 - July 2014 • Interpreted information from a series of database investigations to make predictions and recommendations for a company’s scope of work. • Worked with large datasets storing and processing with NoSQL databases HBase, MongoDB, Microsoft SQL • Discussed results of database analysis with various members of management in an organization, and led staff members to realize the significance of the data. • Discovered industry trends based on data collection methods and analysis strategies, and used the information to help the company make production and product adjustments to increase efficiency by 12 percent. Data Scientist November 2008 - August 2011 • Ensured accurate and consistent statistical analysis by meticulously going through the data and validating results. • Developed company guidelines and best practices based on information learned through an analysis of consumer behavior data. • Determined additional means of organization improvement with employee engagement by using data collected by staff surveys. Education and Training https://www.livecareer.com/build-resume/select-resume Master of Science in Statistics 2008 University of Oregon Eugene, Oregon Bachelor of Science in Mathematics 2004 Stanford University Stanford, California Skills • Expertise in computer programming languages such as Java, Python and C++, and database programs like Hadoop, Hive, Pig Latin, HBase and MongoDB • Excellent skills in analytical analysis and proven ability to read and interpret different points of data • Use of variety data analytics methods and tools such as Machine Learning, clustering, classification, general statistical analysis, data preparation • Strong listening skills, allowing me to carefully consider instructions and feedback from other staff members • Top written and oral communication skills, giving me the necessary tools to be able to present complicated information to a general audience of stakeholders • Outstanding problem-solving method to help me design the best strategies of measuring information and reviewing the results Hobbies and Interests Computer technology enthusiast. Since I was young, I have worked on building computers and putting together hardware in my spare time. Website moderator. I am a volunteer moderator for a popular social media website that is focused on discussing computers and programming. Additionally, I enjoy marathon running every once in a while and pursue daily running goals to maintain my physical endurance.'
EXAMPLE_CV_5 = 'Software Specialist Resume Sample Melody Montoya 123 Main Street, San Francisco, CA 94122 Home: 000-000-0000 | Cell: 000-000-0000 email@example.com Professional Summary Skilled and experienced Software Specialist utilizes strong analytical skills for complex troubleshooting. Extensively knowledgeable about data gathering analysis and sampling techniques.Has a Master’s Degree in Information Technology a Bachelor’s Degree in Computer Information Systems and more than eleven years of Software Specialist experience. Core Qualifications • Excellent written and oral communication skills • Strong analysis and sampling abilities • Extensive knowledge of n-tier architecture methods for troubleshooting including web software COTS and database methods • Good ability to work independently and uses terrific time management skills • Strong resource management and project implementation skills • Above-average problem identification and resolution Experience Software Specialist III 6/1/2009 - 7/1/2014 California Department of Employment and Development New Cityland, CA • Completed complex problem troubleshooting using n-tier architecture methods. • Identified and resolved technological issues. • Installed and supported software and recommended needed hardware. • Kept abreast of new software applications development to stay current. Software Specialist II 5/1/2003 - 5/1/2009 California Department of Employment and Development New Cityland, CA • Proposed and designed information technology solutions to problems and recommended optimal systems changes. • Successfully performed troubleshooting of complex problems and programmed software to meet needs. • Installed systems and maintained them. • Completed all associated documentation and submitted it in a timely manner. Education Master’s of Science Degree - Information Technology 2003 California Pacific University New Cityland, CA Bachelor’s of Sciences Degree - Computer Information Systems 2001 California Pacific University New Cityland, CA'
EXAMPLE_CV_6 = 'DevOps Engineer Laura Baker 55 Capitol Street, Salt Lake City, Utah, 11111 T: 555-123-9485 E: l.baker@anymail Professional Summary Innovative DevOps engineer with diverse work history and educational background. Professional experience as a software developer, systems administrator, and IT support representative. Knowledge of software development, testing, and deployment procedures allows me to bridge the gap between development and operations for an efficient work model. Earned high praise from previous employers including Outstanding SysAdmin Award for managing new software deployment and upgrades. Work Experience Lead Software Developer 2013 to present • Manage software development projects, including checking and improving the code. • Using large scale deployment and cloud automation tools Chef, Puppet, SliepStream • Communicate directly with clients to learn project objectives and create solutions that meet functionality and budget requirements. • Oversee testing, implementation, and upgrade procedures for new software, including giving instructional presentations to clients. Systems Administrator 2011-2013 • Ensured maximum functionality and efficiency for network and other computer systems. • Lead new software deployment for high-level customer as project manager, and earned Outstanding SysAdmin Award. • Managed team of technical support representatives. • Scheduled security patches and planned outages and upgrades to create minimal disruption for employees. • Ensured employees followed security protocols and completed disciplinary action and documentation when necessary. IT Support Representative 2010-2011 • Communicated directly with customers to provide help and troubleshooting advice for a variety of computer and software products. • Documented every problem and solution according to company protocols. • Advised end-users of best practices in computer, network, and software operation. Education and Training Certified Software Development Professional 2012 IEEE Computer Society Online Bachelor of Science in Computer Science 2011 Regis University Colorado Skills • Expert-level proficiency in C++, Adobe ActionScript, Apache Maven, Linux, and Python, Ansible and Chef recipe scripting • Ability to explain complex technical material in an understandable way so software implementation and testing runs efficiently • Extremely organized and detail-oriented • Highly focused with ability to make results-driven decisions Hobbies and Interests Enjoy playing and developing video games and attending gamer conferences. Teach Computer Basics course at local library every month. Moderate an online forum for discussing video games and movies.'

def main():
    print(get_cv_info(EXAMPLE_CV_6))

def get_cv_info(cv):
    job_database = load_jobs(JOB_DATA)
    competence_relevance_database = load_dspp_mapping(COMPETENCE_RELEVANCE)
    jobs = extract_jobs(cv, job_database, competence_relevance_database)
    tagged_jobs = tag_relevant_jobs(jobs)
    return tagged_jobs

'''
Loads the professions / jobs from an excel sheet that has three columns:
First column has job names
The second column has the category the job belongs to
Third column contains mapping to Data science professional profile by EDISON (DSPP)

Input: Path to excel file containing job names, categories, and DSPP
Output: a dictionary containing jobs as keys and dictionary with category and DSPP as value
'''
def load_jobs(path):
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    jobs = {}
    job_titles = ws.__getitem__("A")
    job_categories = ws.__getitem__("B")
    job_DSPP = ws.__getitem__("C")
    for title, category, dspp in zip(job_titles[1:],job_categories[1:], job_DSPP[1:]):
        if title.value:
            if dspp.value:
                jobs[title.value.lower()] = {'category': category.value.lower(), 'DSPP': dspp.value}
            else:
                jobs[title.value.lower()] = {'category': category.value.lower(), 'DSPP': 'NA'}
    return jobs

'''
Loads the mappings from DSPF to DSCF relevance scores
from an excel sheet that has 22 columns (datascience profiles - dspp) and 30 rows with all competences (dscf)

Input: Path to excel file containing relevance scores
Output: a dictionary containing dsp as keys and dictionary with competence and relevance score as value
'''
def load_dspp_mapping(path):
    dspp_dscf_mapping = {}
    # prepare profile labels/keys
    dspp = []
    for i in range(1, 23):
        if i < 10:
            dspp.append('DSP0' + str(i))
        else:
            dspp.append('DSP' + str(i))
        dspp_dscf_mapping[dspp[i - 1]] = {}
    # load competence relevance mappings from file
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    dsp_index = 0
    for column in ws.iter_cols(3, None, 3):
        competence_index = 0
        for cell in column:
            if cell.value:
                mapping_val = str(cell.value)
            else:
                mapping_val = 'NA'
            dspp_dscf_mapping[dspp[dsp_index]].update({competence_list[competence_index]: mapping_val})
            competence_index += 1
        dsp_index += 1
    return dspp_dscf_mapping

'''
Find a list of jobs in a CV
Output: Dictionary with jobs as keys and level, experience, dspp, competence relevance scores as values
'''
def extract_jobs(cv, job_database, competence_relevance_database):
    jobs = {}
    job_index = 0
    # Find jobs
    for job in job_database.keys():
        job_regex = r'[^a-zA-Z]'+job+r'[^a-zA-Z]'
        regular_expression = re.compile(job_regex, re.IGNORECASE)
        regex_result = re.finditer(regular_expression, cv)
        # Keep track of how many times the same job appears in the CV
        same_job_count = 1
        for result in regex_result:
            # Save found job, get job level, get experience years, assign DSPP if applicable
            if same_job_count == 1:
                index = job.capitalize()
            else:
                index = job.capitalize() + '_' + str(same_job_count)
            # save job + get job level
            jobs[index] = extract_position_level(cv, result.start(), result.end())
            # get experience years
            jobs[index].update(extract_experience(cv, result.start(), result.end()))
            # assign dspp
            dspp = job_database[job]['DSPP']
            jobs[index].update({'dspp': dspp})
            # assign competence relevance scores if applicable
            if dspp != 'NA':
                jobs[index].update({'relevance_scores': competence_relevance_database[dspp]})
            else:
                jobs[index].update({'relevance_scores': 'NA'})
            jobs[index].update({'index': job_index})
            job_index += 1
            same_job_count += 1
    return jobs

'''
Returns the level of the position near the text where the job is found in the CV
'''
def extract_position_level(cv, start, end):
    relevant_cv_part = cv[start-100:end+100]
    for level_key in position_level:
        for level in position_level[level_key]:
            level_regex = r'[^a-zA-Z]'+level+r'[^a-zA-Z]'
            regular_expression = re.compile(level_regex, re.IGNORECASE)
            regex_result = re.search(regular_expression, relevant_cv_part)
            if regex_result:
                return {'level': level_key}
    return {'level': 'NA'}

'''
Extracts the amount of experience (years) someone has practised a profession
Input: CV text, start and end position where the job was found
'''
def extract_experience(cv, start, end):
    def get_month_index(month):
        if len(month) > 3:
            month = month[0:3]
        month_dict = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6, 'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}
        return month_dict[month]

    relevant_cv_part = cv[start-100:end+100]
    # Find year duration
    regular_expression = re.compile(year_duration, re.IGNORECASE) 
    regex_result = re.search(regular_expression, relevant_cv_part)
    if regex_result:
        duration = regex_result.group().lower().replace(" ", "")
        if '-' in duration:
            duration_split = duration.split('-')
        elif 'to' in duration:
            duration_split = duration.split('to')
        start = duration_split[0]
        if duration_split[1] == 'now' or duration_split[1] == 'current' or duration_split[1] == 'present':
            end = datetime.datetime.now().year
        else:
            end = duration_split[1]
        experience = calculate_experience(start, end, 'year')
        return {'experience': experience, 'start': start, 'end': end}
    else:
        # Find full date duration
        regular_expression = re.compile(full_date_duration, re.IGNORECASE) 
        regex_result = re.search(regular_expression, relevant_cv_part)
        if regex_result:
            duration = regex_result.group().lower().strip()
            if '-' in duration:
                duration_split = duration.split('-')
            elif 'to' in duration:
                duration_split = duration.split('to')
            if '/' in duration_split[0]:
                start_split = duration_split[0].strip().split('/')
                end_split = duration_split[1].strip().split('/')
            elif '-' in duration_split[0]:
                start_split = duration_split[0].strip().split('-')
                end_split = duration_split[1].strip().split('-')
            start = datetime.date(int(start_split[2]), int(start_split[1]), int(start_split[0]))
            end = datetime.date(int(end_split[2]), int(end_split[1]), int(end_split[0]))
            experience = calculate_experience(start, end, 'full')
            return {'experience': experience, 'start': start.strftime('%d/%m/%y'), 'end': end.strftime('%d/%m/%y')}
        else:
            # Find month year duration
            regular_expression = re.compile(month_year_duration, re.IGNORECASE) 
            regex_result = re.search(regular_expression, relevant_cv_part)
            if regex_result:
                duration = regex_result.group().lower()
                if '-' in duration:
                    duration_split = duration.split('-')
                elif 'to' in duration:
                    duration_split = duration.split('to')
                start_month = get_month_index(duration_split[0].strip().split(' ')[0])
                start = datetime.date(int(duration_split[0].strip().split(' ')[1]), start_month, 1)
                end_month = get_month_index(duration_split[1].strip().split(' ')[0])
                end = datetime.date(int(duration_split[1].strip().split(' ')[1]), end_month, 1)
                experience = calculate_experience(start, end, 'month_year')
                return {'experience': experience, 'start': start.strftime('%d/%m/%y'), 'end': end.strftime('%d/%m/%y')}
    return {'experience': 'NA', 'start': 'NA', 'end': 'NA'}

'''
If a job has values for at least two keys (dspp / experience / level) it can be assumed that it has been tagged correctly
If this is they case, a value will be added to a relevant key to indicate that this job has probably been tagged correctly
'''
def tag_relevant_jobs(jobs):
    for job_key in jobs.keys():
        job = jobs[job_key]
        relevant_values = []
        relevant_values.append(job['dspp'])
        relevant_values.append(job['experience'])
        relevant_values.append(job['level'])
        if relevant_values.count('NA') < 2:
            jobs[job_key].update({'relevant': True})
        else:
            jobs[job_key].update({'relevant': False})
    return jobs

'''
Calculate the experience of someone by using regex to detect month / year patterns
'''
def calculate_experience(start_date, end_date, date_type):
    if date_type == 'year':
        return int(end_date) - int(start_date)
    elif date_type == 'month_year' or date_type == 'full':
        return int((end_date - start_date).days/365)

if __name__ == "__main__":
    main()
